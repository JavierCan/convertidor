from __future__ import annotations

import base64
import io
import json
import re
import time
import zipfile
import zlib
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
import xml.etree.ElementTree as ET

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN
# ============================================================

st.set_page_config(
    page_title="CFDI a Excel",
    page_icon="🧾",
    layout="wide",
)

BASE_DIR = Path(__file__).parent
ASSETS = BASE_DIR / "assets"

PREVIEW_XML_LIMIT = 20
PREVIEW_ROW_LIMIT = 25
EXCEL_MAX_ROWS = 1_048_576
LAYOUT_VERSION = 1

TECHNICAL_FIELDS = {
    "Sello",
    "Certificado",
    "SelloCFD",
    "SelloSAT",
}


# ============================================================
# FRONTEND HTML / CSS / JS
# ============================================================

def load_asset(name: str) -> str:
    path = ASSETS / name
    return path.read_text(encoding="utf-8") if path.exists() else ""


frontend_html = load_asset("index.html")
frontend_html = frontend_html.replace(
    "/*__INLINE_CSS__*/",
    load_asset("styles.css"),
)
frontend_html = frontend_html.replace(
    "/*__INLINE_JS__*/",
    load_asset("script.js"),
)

components.html(
    frontend_html,
    height=300,
    scrolling=False,
)


st.markdown(
    """
    <style>
    .block-container {
        padding-top: 1rem;
        padding-bottom: 3rem;
        max-width: 1500px;
    }

    div[data-testid="stFileUploader"] {
        border: 2px dashed #16a34a;
        border-radius: 18px;
        padding: 10px;
        background: #f0fdf4;
    }

    div.stButton > button,
    div.stDownloadButton > button {
        width: 100%;
        border-radius: 12px;
        font-weight: 700;
        min-height: 46px;
    }

    div.stButton > button[kind="primary"],
    div.stDownloadButton > button {
        background: linear-gradient(135deg, #16a34a, #15803d) !important;
        color: white !important;
        border: none !important;
    }

    div.stButton > button[kind="primary"]:hover,
    div.stDownloadButton > button:hover {
        background: linear-gradient(135deg, #15803d, #166534) !important;
        color: white !important;
    }

    .signature {
        text-align: center;
        color: #64748b;
        font-size: 0.95rem;
        margin-top: 1rem;
        font-style: italic;
    }

    .layout-help {
        padding: 0.85rem 1rem;
        border: 1px solid #bbf7d0;
        background: #f0fdf4;
        border-radius: 12px;
        color: #166534;
        margin-bottom: 0.75rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# UTILIDADES XML
# ============================================================

def local_name(tag: str) -> str:
    """Elimina namespaces del nombre visible de una etiqueta."""
    return tag.split("}")[-1] if "}" in tag else tag.split(":")[-1]


def normalize_col(name: str) -> str:
    name = re.sub(r"\s+", "_", str(name).strip())
    name = re.sub(
        r"[^A-Za-z0-9_áéíóúÁÉÍÓÚñÑ.-]",
        "_",
        name,
    )
    return re.sub(r"_+", "_", name).strip("_")


def identify_xml_type(root: ET.Element) -> str:
    root_name = local_name(root.tag)
    version = (
        root.attrib.get("Version")
        or root.attrib.get("version")
        or ""
    )

    names = {
        local_name(element.tag)
        for element in root.iter()
    }

    if root_name == "Comprobante":
        payment_nodes = {
            "Pagos",
            "Pago",
            "DoctoRelacionado",
        }

        if payment_nodes & names:
            return (
                f"CFDI {version or 'sin versión'} "
                "- Complemento de pago"
            )

        return f"CFDI {version or 'sin versión'}"

    return f"XML genérico - {root_name}"


def collect_flat(
    element: ET.Element,
    prefix: str = "",
    skip_tags: set[str] | None = None,
    exclude_technical: bool = True,
) -> dict[str, str]:
    """
    Aplana atributos y textos del XML de forma dinámica.

    No depende de nombres de proveedor, UUID o campos concretos.
    """
    skip_tags = skip_tags or set()
    result: dict[str, str] = {}

    element_name = local_name(element.tag)

    if element_name in skip_tags:
        return result

    current_path = normalize_col(
        f"{prefix}_{element_name}"
        if prefix
        else element_name
    )

    for attribute, value in element.attrib.items():
        attribute_name = local_name(attribute)

        if (
            exclude_technical
            and attribute_name in TECHNICAL_FIELDS
        ):
            continue

        column_name = normalize_col(
            f"{current_path}_{attribute_name}"
        )

        candidate = column_name
        suffix = 2

        while (
            candidate in result
            and result[candidate] != value
        ):
            candidate = f"{column_name}_{suffix}"
            suffix += 1

        result[candidate] = value

    element_text = (element.text or "").strip()

    if element_text and len(element) == 0:
        result[
            normalize_col(f"{current_path}_Texto")
        ] = element_text

    child_counts = Counter(
        local_name(child.tag)
        for child in element
    )

    child_occurrences: defaultdict[str, int] = defaultdict(int)

    for child in element:
        child_name = local_name(child.tag)
        child_occurrences[child_name] += 1

        child_prefix = current_path

        if child_counts[child_name] > 1:
            child_prefix = normalize_col(
                f"{current_path}_{child_name}_"
                f"{child_occurrences[child_name]}"
            )
            child_prefix = child_prefix.rsplit("_", 1)[0]

        child_data = collect_flat(
            child,
            prefix=child_prefix,
            skip_tags=skip_tags,
            exclude_technical=exclude_technical,
        )

        for key, value in child_data.items():
            candidate = key
            suffix = 2

            while (
                candidate in result
                and result[candidate] != value
            ):
                candidate = f"{key}_{suffix}"
                suffix += 1

            result[candidate] = value

    return result


def find_elements(
    root: ET.Element,
    tag_name: str,
) -> list[ET.Element]:
    return [
        element
        for element in root.iter()
        if local_name(element.tag) == tag_name
    ]


def parse_xml(
    raw: bytes,
    source_name: str,
    representation: str,
    exclude_technical: bool,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        root = ET.fromstring(raw)
    except ET.ParseError as exc:
        raise ValueError(f"XML inválido: {exc}") from exc

    xml_type = identify_xml_type(root)
    concepts = find_elements(root, "Concepto")

    base_row: dict[str, Any] = {
        "Archivo_XML": source_name,
        "Tipo_XML": xml_type,
    }

    base_row.update(
        collect_flat(
            root,
            skip_tags={"Conceptos"},
            exclude_technical=exclude_technical,
        )
    )

    metadata = {
        "tipo": xml_type,
        "conceptos": len(concepts),
    }

    if representation == "por_comprobante":
        document_row = dict(base_row)
        document_row["Total_Conceptos"] = len(concepts)

        for index, concept in enumerate(
            concepts,
            start=1,
        ):
            concept_data = collect_flat(
                concept,
                prefix=f"Concepto_{index}",
                exclude_technical=exclude_technical,
            )
            document_row.update(concept_data)

        return [document_row], metadata

    if not concepts:
        return [base_row], metadata

    rows: list[dict[str, Any]] = []

    for index, concept in enumerate(
        concepts,
        start=1,
    ):
        row = dict(base_row)
        row["Concepto_Indice"] = index
        row.update(
            collect_flat(
                concept,
                exclude_technical=exclude_technical,
            )
        )
        rows.append(row)

    return rows, metadata


# ============================================================
# ARCHIVOS DE ENTRADA
# ============================================================

def iter_xmls(
    uploaded_files,
) -> Iterable[tuple[str, bytes]]:
    """
    Recorre XML sueltos y XML contenidos dentro de ZIP.
    """
    for uploaded in uploaded_files:
        file_name = uploaded.name
        raw = uploaded.getvalue()
        suffix = Path(file_name).suffix.lower()

        if suffix == ".xml":
            yield file_name, raw

        elif suffix == ".zip":
            try:
                with zipfile.ZipFile(
                    io.BytesIO(raw),
                    "r",
                ) as zip_file:
                    for member in zip_file.infolist():
                        if member.is_dir():
                            continue

                        if (
                            Path(member.filename).suffix.lower()
                            == ".xml"
                        ):
                            yield (
                                f"{Path(file_name).stem}/"
                                f"{member.filename}",
                                zip_file.read(member),
                            )

            except zipfile.BadZipFile as exc:
                raise ValueError(
                    f"{file_name}: ZIP inválido"
                ) from exc


def analyze(uploaded_files) -> dict[str, Any]:
    total = 0
    valid = 0
    concepts = 0
    types: Counter[str] = Counter()
    errors: list[dict[str, str]] = []

    for file_name, raw in iter_xmls(uploaded_files):
        total += 1

        try:
            root = ET.fromstring(raw)
            valid += 1
            types[identify_xml_type(root)] += 1
            concepts += len(
                find_elements(root, "Concepto")
            )

        except Exception as exc:
            errors.append(
                {
                    "Archivo": file_name,
                    "Error": str(exc),
                }
            )

    return {
        "total": total,
        "valid": valid,
        "concepts": concepts,
        "types": dict(types),
        "errors": errors,
    }


def make_preview(
    uploaded_files,
    representation: str,
    exclude_technical: bool,
    sample_mode: str,
) -> tuple[
    pd.DataFrame,
    list[dict[str, str]],
    int,
]:
    xml_files = list(iter_xmls(uploaded_files))

    if (
        sample_mode == "Aleatorios"
        and len(xml_files) > PREVIEW_XML_LIMIT
    ):
        sample = (
            pd.Series(xml_files)
            .sample(
                PREVIEW_XML_LIMIT,
                random_state=42,
            )
            .tolist()
        )
    else:
        sample = xml_files[:PREVIEW_XML_LIMIT]

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []

    for file_name, raw in sample:
        try:
            parsed_rows, _ = parse_xml(
                raw,
                file_name,
                representation,
                exclude_technical,
            )
            rows.extend(parsed_rows)

        except Exception as exc:
            errors.append(
                {
                    "Archivo": file_name,
                    "Error": str(exc),
                }
            )

    return pd.DataFrame(rows), errors, len(sample)


# ============================================================
# LAYOUTS DE COLUMNAS
# ============================================================

def encode_layout(payload: dict[str, Any]) -> str:
    raw = json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")

    compressed = zlib.compress(raw, level=9)

    return (
        base64.urlsafe_b64encode(compressed)
        .decode("ascii")
        .rstrip("=")
    )


def decode_layout(token: str) -> dict[str, Any]:
    padding = "=" * (-len(token) % 4)
    compressed = base64.urlsafe_b64decode(
        token + padding
    )
    raw = zlib.decompress(compressed)
    payload = json.loads(raw.decode("utf-8"))

    if not isinstance(payload, dict):
        raise ValueError("El layout no tiene un formato válido.")

    return payload


def validate_layout(
    payload: dict[str, Any],
) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise ValueError("El layout debe ser un objeto JSON.")

    representation = payload.get(
        "representation",
        "por_concepto",
    )

    if representation not in {
        "por_concepto",
        "por_comprobante",
    }:
        raise ValueError(
            "La representación del layout no es válida."
        )

    all_columns = payload.get(
        "all_columns_order",
        [],
    )
    selected_columns = payload.get(
        "selected_columns",
        [],
    )

    if not isinstance(all_columns, list):
        raise ValueError(
            "`all_columns_order` debe ser una lista."
        )

    if not isinstance(selected_columns, list):
        raise ValueError(
            "`selected_columns` debe ser una lista."
        )

    payload["version"] = int(
        payload.get("version", LAYOUT_VERSION)
    )
    payload["representation"] = representation
    payload["all_columns_order"] = [
        str(column)
        for column in all_columns
    ]
    payload["selected_columns"] = [
        str(column)
        for column in selected_columns
    ]
    payload["include_new_columns"] = bool(
        payload.get("include_new_columns", True)
    )
    payload["exclude_technical"] = bool(
        payload.get("exclude_technical", True)
    )
    payload["name"] = str(
        payload.get("name", "Mi layout")
    ).strip() or "Mi layout"

    return payload


def current_query_layout() -> (
    tuple[str | None, dict[str, Any] | None, str | None]
):
    token = st.query_params.get("layout")

    if isinstance(token, list):
        token = token[0] if token else None

    if not token:
        return None, None, None

    try:
        return token, validate_layout(
            decode_layout(token)
        ), None
    except Exception as exc:
        return token, None, str(exc)


def build_layout_table(
    available_columns: list[str],
    saved_layout: dict[str, Any] | None,
    representation: str,
) -> pd.DataFrame:
    saved_applies = bool(
        saved_layout
        and saved_layout.get("representation")
        == representation
    )

    if saved_applies:
        saved_order = list(
            dict.fromkeys(
                saved_layout.get(
                    "all_columns_order",
                    [],
                )
            )
        )
        saved_selected = set(
            saved_layout.get(
                "selected_columns",
                [],
            )
        )
    else:
        saved_order = []
        saved_selected = set()

    combined_columns = list(saved_order)

    for column in available_columns:
        if column not in combined_columns:
            combined_columns.append(column)

    rows: list[dict[str, Any]] = []

    for index, column in enumerate(
        combined_columns,
        start=1,
    ):
        is_new_column = column not in saved_order

        if saved_applies:
            include = (
                column in saved_selected
                if not is_new_column
                else bool(
                    saved_layout.get(
                        "include_new_columns",
                        True,
                    )
                )
            )
        else:
            include = True

        rows.append(
            {
                "Incluir": include,
                "Orden": index,
                "Columna": column,
                "Detectada_en_muestra": (
                    column in available_columns
                ),
            }
        )

    return pd.DataFrame(rows)


def selected_columns_from_editor(
    editor_data: pd.DataFrame,
) -> tuple[list[str], list[str]]:
    if editor_data.empty:
        return [], []

    cleaned = editor_data.copy()
    cleaned["Incluir"] = (
        cleaned["Incluir"]
        .fillna(False)
        .astype(bool)
    )
    cleaned["Orden"] = pd.to_numeric(
        cleaned["Orden"],
        errors="coerce",
    ).fillna(999_999)

    cleaned["_posicion_original"] = range(
        len(cleaned)
    )

    cleaned = cleaned.sort_values(
        [
            "Orden",
            "_posicion_original",
        ],
        kind="stable",
    )

    all_columns_order = (
        cleaned["Columna"]
        .astype(str)
        .tolist()
    )

    selected_columns = (
        cleaned.loc[
            cleaned["Incluir"],
            "Columna",
        ]
        .astype(str)
        .tolist()
    )

    return selected_columns, all_columns_order


def make_layout_payload(
    layout_name: str,
    representation: str,
    selected_columns: list[str],
    all_columns_order: list[str],
    include_new_columns: bool,
    exclude_technical: bool,
    output_mode: str,
) -> dict[str, Any]:
    return {
        "version": LAYOUT_VERSION,
        "name": (
            layout_name.strip()
            or "Mi layout"
        ),
        "saved_at_utc": datetime.now(
            timezone.utc
        ).isoformat(),
        "representation": representation,
        "selected_columns": selected_columns,
        "all_columns_order": all_columns_order,
        "include_new_columns": (
            include_new_columns
        ),
        "exclude_technical": exclude_technical,
        "output_mode": output_mode,
    }


# ============================================================
# EXPORTACIÓN
# ============================================================

def sanitize_sheet(name: str) -> str:
    return (
        re.sub(
            r"[\[\]\*\?/\\:]",
            "_",
            name,
        )[:31]
        or "Hoja"
    ).strip()


def format_sheet(
    worksheet,
    headers: list[str],
) -> None:
    fill = PatternFill(
        "solid",
        fgColor="166534",
    )
    font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for index, header in enumerate(
        headers,
        start=1,
    ):
        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = min(
            max(len(str(header)) + 2, 12),
            40,
        )


def resolve_export_columns(
    rows: list[dict[str, Any]],
    selected_columns: list[str] | None,
    include_new_columns: bool,
) -> list[str]:
    discovered_columns: list[str] = []
    discovered_set: set[str] = set()

    for row in rows:
        for column in row:
            if column not in discovered_set:
                discovered_set.add(column)
                discovered_columns.append(column)

    if selected_columns is None:
        return discovered_columns

    columns = list(
        dict.fromkeys(selected_columns)
    )

    if include_new_columns:
        for column in discovered_columns:
            if column not in columns:
                columns.append(column)

    return columns


def rows_to_excel(
    rows: list[dict[str, Any]],
    sheet_name: str = "CFDI_Consolidado",
    errors: list[dict[str, str]] | None = None,
    selected_columns: list[str] | None = None,
    include_new_columns: bool = True,
) -> bytes:
    columns = resolve_export_columns(
        rows,
        selected_columns,
        include_new_columns,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    if not rows:
        worksheet = workbook.create_sheet(
            "Sin_datos"
        )

        if columns:
            worksheet.append(columns)
            format_sheet(
                worksheet,
                columns,
            )
        else:
            worksheet.append(["Mensaje"])
            worksheet.append(
                ["No se generaron filas"]
            )

    else:
        worksheet = None
        row_count = 0
        sheet_index = 1

        for row in rows:
            if (
                worksheet is None
                or row_count
                >= EXCEL_MAX_ROWS - 1
            ):
                current_name = (
                    sheet_name
                    if sheet_index == 1
                    else f"{sheet_name}_{sheet_index}"
                )

                worksheet = workbook.create_sheet(
                    sanitize_sheet(current_name)
                )
                worksheet.append(columns)
                format_sheet(
                    worksheet,
                    columns,
                )
                row_count = 0
                sheet_index += 1

            worksheet.append(
                [
                    row.get(column, "")
                    for column in columns
                ]
            )
            row_count += 1

    if errors:
        error_sheet = workbook.create_sheet(
            "Errores"
        )
        error_headers = [
            "Archivo",
            "Error",
        ]
        error_sheet.append(error_headers)

        for error in errors:
            error_sheet.append(
                [
                    error.get("Archivo", ""),
                    error.get("Error", ""),
                ]
            )

        format_sheet(
            error_sheet,
            error_headers,
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()


def process_consolidated(
    uploaded_files,
    representation: str,
    exclude_technical: bool,
    progress,
    status,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, str]],
]:
    xml_files = list(iter_xmls(uploaded_files))
    rows: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []

    for index, (file_name, raw) in enumerate(
        xml_files,
        start=1,
    ):
        try:
            parsed_rows, _ = parse_xml(
                raw,
                file_name,
                representation,
                exclude_technical,
            )
            rows.extend(parsed_rows)

        except Exception as exc:
            errors.append(
                {
                    "Archivo": file_name,
                    "Error": str(exc),
                }
            )

        progress.progress(
            index / len(xml_files)
        )
        status.caption(
            f"Procesando {index:,} "
            f"de {len(xml_files):,}: "
            f"{file_name}"
        )

    return rows, errors


def process_individual(
    uploaded_files,
    representation: str,
    exclude_technical: bool,
    selected_columns: list[str],
    include_new_columns: bool,
    progress,
    status,
) -> tuple[
    bytes,
    list[dict[str, str]],
]:
    xml_files = list(iter_xmls(uploaded_files))
    output = io.BytesIO()
    errors: list[dict[str, str]] = []

    used_names: Counter[str] = Counter()

    with zipfile.ZipFile(
        output,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as zip_output:
        for index, (file_name, raw) in enumerate(
            xml_files,
            start=1,
        ):
            try:
                rows, _ = parse_xml(
                    raw,
                    file_name,
                    representation,
                    exclude_technical,
                )

                base_name = (
                    f"{Path(file_name).stem}"
                    "_convertido.xlsx"
                )

                used_names[base_name] += 1

                if used_names[base_name] > 1:
                    stem = Path(base_name).stem
                    output_name = (
                        f"{stem}_"
                        f"{used_names[base_name]}.xlsx"
                    )
                else:
                    output_name = base_name

                excel_bytes = rows_to_excel(
                    rows,
                    "CFDI",
                    selected_columns=selected_columns,
                    include_new_columns=(
                        include_new_columns
                    ),
                )

                zip_output.writestr(
                    output_name,
                    excel_bytes,
                )

            except Exception as exc:
                errors.append(
                    {
                        "Archivo": file_name,
                        "Error": str(exc),
                    }
                )

            progress.progress(
                index / len(xml_files)
            )
            status.caption(
                f"Procesando {index:,} "
                f"de {len(xml_files):,}: "
                f"{file_name}"
            )

        if errors:
            error_csv = (
                pd.DataFrame(errors)
                .to_csv(index=False)
                .encode("utf-8-sig")
            )

            zip_output.writestr(
                "reporte_errores.csv",
                error_csv,
            )

    output.seek(0)

    return output.getvalue(), errors


# ============================================================
# ESTADO DE SESIÓN Y LAYOUT DE URL
# ============================================================

SESSION_DEFAULTS = {
    "download_data": None,
    "download_name": None,
    "download_mime": None,
    "summary": None,
    "active_layout": None,
    "loaded_query_token": None,
    "layout_editor_version": 0,
    "layout_seed_by_representation": {},
    "layout_upload_signature": None,
}

for session_key, default_value in (
    SESSION_DEFAULTS.items()
):
    if session_key not in st.session_state:
        st.session_state[session_key] = (
            default_value.copy()
            if isinstance(default_value, dict)
            else default_value
        )


query_token, query_layout, query_error = (
    current_query_layout()
)

if (
    query_token
    and query_token
    != st.session_state.loaded_query_token
):
    st.session_state.loaded_query_token = (
        query_token
    )

    if query_layout:
        st.session_state.active_layout = (
            query_layout
        )
        st.session_state[
            "exclude_technical_widget"
        ] = query_layout.get(
            "exclude_technical",
            True,
        )
        st.session_state[
            "representation_main"
        ] = (
            "Una fila por concepto"
            if query_layout.get(
                "representation"
            )
            == "por_concepto"
            else "Una fila por comprobante"
        )


active_layout = st.session_state.active_layout

if "exclude_technical_widget" not in (
    st.session_state
):
    st.session_state[
        "exclude_technical_widget"
    ] = bool(
        active_layout.get(
            "exclude_technical",
            True,
        )
        if active_layout
        else True
    )


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:
    st.header("⚙️ Opciones avanzadas")

    exclude_technical = st.checkbox(
        "Excluir sellos y certificados largos",
        key="exclude_technical_widget",
        help=(
            "Excluye Sello, Certificado, SelloCFD "
            "y SelloSAT para reducir el tamaño."
        ),
    )

    sample_mode = st.radio(
        "Muestra del preview",
        [
            "Primeros archivos",
            "Aleatorios",
        ],
        index=0,
        help=(
            "Con cargas grandes, el preview "
            "utiliza hasta 20 XML."
        ),
    )

    st.markdown("---")

    if query_error:
        st.warning(
            "El layout de la URL no pudo cargarse: "
            f"{query_error}"
        )

    if active_layout:
        st.success(
            "Layout activo: "
            f"**{active_layout.get('name', 'Mi layout')}**"
        )

    st.caption(
        "El layout puede guardarse en la URL "
        "y como archivo JSON."
    )


# ============================================================
# 1. CARGA
# ============================================================

st.markdown("## 1. Carga tus archivos")

st.caption(
    "Puedes cargar un XML, varios XML, "
    "un ZIP con XML o varios ZIP."
)

uploaded_files = st.file_uploader(
    "Arrastra o selecciona XML y ZIP",
    type=["xml", "zip"],
    accept_multiple_files=True,
    help=(
        "Admite uno o múltiples archivos "
        "XML y ZIP."
    ),
)

if not uploaded_files:
    st.info(
        "Carga al menos un XML o ZIP "
        "para comenzar."
    )
    st.stop()

try:
    analysis = analyze(uploaded_files)
except Exception as exc:
    st.error(
        f"No se pudo analizar la carga: {exc}"
    )
    st.stop()

if analysis["total"] == 0:
    st.warning(
        "No se encontraron archivos XML "
        "dentro de la carga."
    )
    st.stop()


input_signature = tuple(
    sorted(
        (
            uploaded.name,
            int(
                getattr(
                    uploaded,
                    "size",
                    len(uploaded.getvalue()),
                )
            ),
        )
        for uploaded in uploaded_files
    )
)

if (
    st.session_state.get("input_signature")
    != input_signature
):
    st.session_state.input_signature = (
        input_signature
    )
    st.session_state.download_data = None
    st.session_state.download_name = None
    st.session_state.download_mime = None
    st.session_state.summary = None
    st.session_state.layout_editor_version += 1
    st.session_state.layout_seed_by_representation = {}


# ============================================================
# 2. ANÁLISIS
# ============================================================

st.markdown("## 2. Análisis automático")

analysis_columns = st.columns(4)

analysis_columns[0].metric(
    "XML detectados",
    f"{analysis['total']:,}",
)
analysis_columns[1].metric(
    "XML válidos",
    f"{analysis['valid']:,}",
)
analysis_columns[2].metric(
    "Conceptos detectados",
    f"{analysis['concepts']:,}",
)
analysis_columns[3].metric(
    "Tipos de estructura",
    f"{len(analysis['types']):,}",
)

if analysis["types"]:
    st.dataframe(
        pd.DataFrame(
            [
                {
                    "Tipo detectado": xml_type,
                    "Cantidad": quantity,
                }
                for xml_type, quantity
                in analysis["types"].items()
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )

if analysis["errors"]:
    with st.expander(
        f"⚠️ Ver "
        f"{len(analysis['errors']):,} "
        "XML con problemas"
    ):
        st.dataframe(
            pd.DataFrame(
                analysis["errors"]
            ),
            use_container_width=True,
            hide_index=True,
        )


# ============================================================
# 3. CARGAR LAYOUT GUARDADO
# ============================================================

st.markdown("## 3. Carga o reutiliza un layout")

layout_upload = st.file_uploader(
    "Cargar layout guardado (.json)",
    type=["json"],
    accept_multiple_files=False,
    key="layout_json_uploader",
    help=(
        "El JSON conserva la estructura, "
        "columnas seleccionadas y orden."
    ),
)

if layout_upload is not None:
    layout_bytes = layout_upload.getvalue()
    upload_signature = (
        layout_upload.name,
        len(layout_bytes),
        hash(layout_bytes),
    )

    if (
        upload_signature
        != st.session_state.layout_upload_signature
    ):
        try:
            loaded_layout = validate_layout(
                json.loads(
                    layout_bytes.decode("utf-8")
                )
            )

            st.session_state.active_layout = (
                loaded_layout
            )
            st.session_state.layout_upload_signature = (
                upload_signature
            )
            st.session_state[
                "exclude_technical_widget"
            ] = loaded_layout.get(
                "exclude_technical",
                True,
            )
            st.session_state[
                "representation_main"
            ] = (
                "Una fila por concepto"
                if loaded_layout.get(
                    "representation"
                )
                == "por_concepto"
                else "Una fila por comprobante"
            )
            st.session_state.layout_editor_version += 1
            st.session_state.layout_seed_by_representation = {}

            st.rerun()

        except Exception as exc:
            st.error(
                "No fue posible cargar el layout: "
                f"{exc}"
            )

active_layout = st.session_state.active_layout

if active_layout:
    st.success(
        "Se aplicará el layout "
        f"**{active_layout.get('name', 'Mi layout')}** "
        "a las columnas compatibles."
    )
else:
    st.caption(
        "No hay un layout guardado activo. "
        "La app seleccionará inicialmente "
        "todas las columnas detectadas."
    )


# ============================================================
# 4. ESTRUCTURA DE FILAS
# ============================================================

st.markdown("## 4. Elige la estructura de las filas")

if "representation_main" not in (
    st.session_state
):
    st.session_state.representation_main = (
        "Una fila por concepto"
    )

representation_label = st.radio(
    "Estructura del Excel",
    [
        "Una fila por concepto",
        "Una fila por comprobante",
    ],
    horizontal=True,
    key="representation_main",
)

representation = (
    "por_concepto"
    if representation_label
    == "Una fila por concepto"
    else "por_comprobante"
)

description_columns = st.columns(2)

with description_columns[0]:
    with st.container(border=True):
        st.markdown(
            "### 🧾 Una fila por concepto"
        )
        st.write(
            "Cada producto o servicio genera "
            "una fila."
        )
        st.caption(
            "Recomendado para análisis detallado "
            "y Power BI."
        )

with description_columns[1]:
    with st.container(border=True):
        st.markdown(
            "### 📄 Una fila por comprobante"
        )
        st.write(
            "Cada XML genera una sola fila."
        )
        st.caption(
            "Recomendado para revisar facturas "
            "y totales generales."
        )


# ============================================================
# 5. PREVIEW Y DISEÑO DE COLUMNAS
# ============================================================

st.markdown("## 5. Diseña las columnas del Excel")

with st.spinner(
    "Detectando columnas y preparando preview..."
):
    preview_df, preview_errors, sample_count = (
        make_preview(
            uploaded_files,
            representation,
            exclude_technical,
            sample_mode,
        )
    )

preview_metrics = st.columns(3)
preview_metrics[0].metric(
    "XML usados en muestra",
    f"{sample_count:,}",
)
preview_metrics[1].metric(
    "Filas en preview",
    f"{len(preview_df):,}",
)
preview_metrics[2].metric(
    "Columnas detectadas",
    f"{len(preview_df.columns):,}",
)

available_columns = list(preview_df.columns)

saved_layout_for_view = (
    active_layout
    if (
        active_layout
        and active_layout.get(
            "representation"
        )
        == representation
    )
    else None
)

seed_data = (
    st.session_state
    .layout_seed_by_representation
    .get(representation)
)

if seed_data:
    initial_layout_table = pd.DataFrame(
        seed_data
    )
else:
    initial_layout_table = build_layout_table(
        available_columns,
        saved_layout_for_view,
        representation,
    )

editor_key = (
    f"column_layout_editor_"
    f"{representation}_"
    f"{st.session_state.layout_editor_version}"
)

st.markdown(
    """
    <div class="layout-help">
    <strong>Cómo mover columnas:</strong>
    cambia el número de <strong>Orden</strong>.
    La columna con orden 1 será la primera del Excel.
    Desmarca <strong>Incluir</strong> para no descargar una columna.
    </div>
    """,
    unsafe_allow_html=True,
)

edited_layout = st.data_editor(
    initial_layout_table,
    key=editor_key,
    use_container_width=True,
    hide_index=True,
    num_rows="fixed",
    disabled=[
        "Columna",
        "Detectada_en_muestra",
    ],
    column_config={
        "Incluir": st.column_config.CheckboxColumn(
            "Incluir",
            help=(
                "Activa o desactiva la columna "
                "en el archivo final."
            ),
            default=True,
        ),
        "Orden": st.column_config.NumberColumn(
            "Orden",
            help=(
                "Número menor = aparece primero."
            ),
            min_value=1,
            step=1,
            format="%d",
        ),
        "Columna": st.column_config.TextColumn(
            "Nombre de columna",
            width="large",
        ),
        "Detectada_en_muestra": (
            st.column_config.CheckboxColumn(
                "Detectada",
                help=(
                    "Indica si apareció en "
                    "la muestra actual."
                ),
            )
        ),
    },
)

selected_columns, all_columns_order = (
    selected_columns_from_editor(
        edited_layout
    )
)

action_columns = st.columns(3)

if action_columns[0].button(
    "Seleccionar todas",
    use_container_width=True,
):
    updated = edited_layout.copy()
    updated["Incluir"] = True
    st.session_state.layout_seed_by_representation[
        representation
    ] = updated.to_dict("records")
    st.session_state.layout_editor_version += 1
    st.rerun()

if action_columns[1].button(
    "Quitar todas",
    use_container_width=True,
):
    updated = edited_layout.copy()
    updated["Incluir"] = False
    st.session_state.layout_seed_by_representation[
        representation
    ] = updated.to_dict("records")
    st.session_state.layout_editor_version += 1
    st.rerun()

if action_columns[2].button(
    "Restablecer orden",
    use_container_width=True,
):
    reset_table = build_layout_table(
        available_columns,
        None,
        representation,
    )
    st.session_state.layout_seed_by_representation[
        representation
    ] = reset_table.to_dict("records")
    st.session_state.layout_editor_version += 1
    st.rerun()


default_include_new = (
    bool(
        saved_layout_for_view.get(
            "include_new_columns",
            True,
        )
    )
    if saved_layout_for_view
    else True
)

include_new_columns = st.toggle(
    "Agregar al final las columnas nuevas "
    "que aparezcan al procesar todos los XML",
    value=default_include_new,
    key=(
        f"include_new_columns_"
        f"{representation}"
    ),
    help=(
        "El preview usa una muestra. "
        "Si otros XML contienen campos nuevos, "
        "puedes agregarlos automáticamente "
        "después del layout elegido."
    ),
)

if not selected_columns:
    st.error(
        "Selecciona al menos una columna "
        "para habilitar la exportación."
    )
else:
    st.success(
        f"Se descargarán "
        f"**{len(selected_columns):,} columnas** "
        "en el orden configurado."
    )

if (
    not edited_layout.empty
    and edited_layout.loc[
        edited_layout["Incluir"].fillna(False),
        "Orden",
    ].duplicated().any()
):
    st.warning(
        "Hay números de orden repetidos. "
        "La app conservará el orden actual "
        "para resolver los empates."
    )


st.markdown("### Vista previa del layout final")

if preview_df.empty:
    st.warning(
        "No fue posible generar filas "
        "para la vista previa."
    )
elif selected_columns:
    preview_layout_df = preview_df.reindex(
        columns=selected_columns
    )

    st.dataframe(
        preview_layout_df.head(
            PREVIEW_ROW_LIMIT
        ),
        use_container_width=True,
        hide_index=True,
        height=430,
    )

    st.caption(
        f"Mostrando hasta "
        f"{PREVIEW_ROW_LIMIT} filas. "
        "El Excel respetará la selección "
        "y el orden mostrados."
    )

if preview_errors:
    with st.expander(
        f"⚠️ Errores del preview: "
        f"{len(preview_errors):,}"
    ):
        st.dataframe(
            pd.DataFrame(preview_errors),
            use_container_width=True,
            hide_index=True,
        )


# ============================================================
# 6. TIPO DE DESCARGA
# ============================================================

st.markdown("## 6. Elige el tipo de descarga")

if analysis["total"] == 1:
    output_mode = "individual"

    st.info(
        "Se detectó un solo XML. "
        "Se generará directamente un Excel "
        "individual; no hay archivos que unir."
    )

else:
    default_output_label = (
        "Unir todos en un solo Excel"
    )

    if (
        active_layout
        and active_layout.get("output_mode")
        == "individual"
    ):
        default_output_label = (
            "Dejar un Excel por cada XML"
        )

    if "output_mode_main" not in (
        st.session_state
    ):
        st.session_state.output_mode_main = (
            default_output_label
        )

    output_mode_label = st.radio(
        "¿Cómo deseas descargar los XML?",
        [
            "Unir todos en un solo Excel",
            "Dejar un Excel por cada XML",
        ],
        horizontal=True,
        key="output_mode_main",
    )

    output_mode = (
        "consolidated"
        if output_mode_label
        == "Unir todos en un solo Excel"
        else "individual"
    )

    if output_mode == "consolidated":
        st.success(
            "Todos los XML se integrarán "
            "en una sola tabla."
        )
    else:
        st.info(
            "Se descargará un ZIP con "
            "un Excel por cada XML."
        )


# ============================================================
# 7. GUARDAR LAYOUT
# ============================================================

st.markdown("## 7. Guarda esta vista para reutilizarla")

layout_name_default = (
    active_layout.get("name", "Mi layout CFDI")
    if active_layout
    else "Mi layout CFDI"
)

layout_name = st.text_input(
    "Nombre del layout",
    value=layout_name_default,
    help=(
        "Ejemplo: Compras mensuales, "
        "Facturas por concepto, Resumen CFDI."
    ),
)

current_layout_payload = make_layout_payload(
    layout_name=layout_name,
    representation=representation,
    selected_columns=selected_columns,
    all_columns_order=all_columns_order,
    include_new_columns=include_new_columns,
    exclude_technical=exclude_technical,
    output_mode=output_mode,
)

layout_json_bytes = json.dumps(
    current_layout_payload,
    ensure_ascii=False,
    indent=2,
).encode("utf-8")

layout_buttons = st.columns(3)

if layout_buttons[0].button(
    "Guardar layout en esta URL",
    type="primary",
    use_container_width=True,
    disabled=not selected_columns,
):
    layout_token = encode_layout(
        current_layout_payload
    )

    st.query_params["layout"] = layout_token
    st.session_state.active_layout = (
        current_layout_payload
    )
    st.session_state.loaded_query_token = (
        layout_token
    )

    st.success(
        "Layout guardado en la URL actual. "
        "Guarda esta página en favoritos "
        "y la próxima vez se seleccionarán "
        "automáticamente las mismas columnas."
    )

layout_buttons[1].download_button(
    "Descargar layout JSON",
    data=layout_json_bytes,
    file_name=(
        f"{normalize_col(layout_name) or 'layout_cfdi'}.json"
    ),
    mime="application/json",
    use_container_width=True,
    disabled=not selected_columns,
)

if layout_buttons[2].button(
    "Borrar layout guardado",
    use_container_width=True,
):
    if "layout" in st.query_params:
        del st.query_params["layout"]

    st.session_state.active_layout = None
    st.session_state.loaded_query_token = None
    st.session_state.layout_editor_version += 1
    st.session_state.layout_seed_by_representation = {}

    st.rerun()

st.caption(
    "El guardado en URL funciona aunque "
    "la aplicación se reinicie, siempre que "
    "abras la misma URL o la tengas en favoritos. "
    "El JSON sirve como respaldo portátil."
)


# ============================================================
# 8. PROCESAR Y DESCARGAR
# ============================================================

st.markdown("## 8. Procesar y descargar")

summary_columns = st.columns(4)

summary_columns[0].metric(
    "XML por procesar",
    f"{analysis['total']:,}",
)
summary_columns[1].metric(
    "Estructura",
    representation_label,
)
summary_columns[2].metric(
    "Columnas",
    f"{len(selected_columns):,}",
)
summary_columns[3].metric(
    "Salida",
    (
        "Excel individual"
        if analysis["total"] == 1
        else (
            "Excel consolidado"
            if output_mode == "consolidated"
            else "ZIP de Excel"
        )
    ),
)

process_label = (
    "Convertir este XML a Excel"
    if analysis["total"] == 1
    else f"Procesar {analysis['total']:,} XML"
)

if st.button(
    process_label,
    type="primary",
    use_container_width=True,
    disabled=not selected_columns,
):
    st.session_state.download_data = None
    st.session_state.download_name = None
    st.session_state.download_mime = None
    st.session_state.summary = None

    progress = st.progress(0)
    status = st.empty()
    started = time.time()

    try:
        if (
            analysis["total"] == 1
            or output_mode == "consolidated"
        ):
            rows, errors = process_consolidated(
                uploaded_files,
                representation,
                exclude_technical,
                progress,
                status,
            )

            excel_bytes = rows_to_excel(
                rows,
                (
                    "CFDI"
                    if analysis["total"] == 1
                    else "CFDI_Consolidado"
                ),
                errors=errors,
                selected_columns=selected_columns,
                include_new_columns=(
                    include_new_columns
                ),
            )

            st.session_state.download_data = (
                excel_bytes
            )

            if analysis["total"] == 1:
                first_name = next(
                    iter_xmls(uploaded_files)
                )[0]
                st.session_state.download_name = (
                    f"{Path(first_name).stem}"
                    "_convertido.xlsx"
                )
            else:
                st.session_state.download_name = (
                    "CFDI_Consolidado.xlsx"
                )

            st.session_state.download_mime = (
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )

            generated_rows = len(rows)

        else:
            zip_bytes, errors = process_individual(
                uploaded_files,
                representation,
                exclude_technical,
                selected_columns,
                include_new_columns,
                progress,
                status,
            )

            st.session_state.download_data = (
                zip_bytes
            )
            st.session_state.download_name = (
                "CFDI_Individuales.zip"
            )
            st.session_state.download_mime = (
                "application/zip"
            )
            generated_rows = None

        st.session_state.summary = {
            "procesados": analysis["total"],
            "correctos": (
                analysis["total"] - len(errors)
            ),
            "errores": len(errors),
            "filas": generated_rows,
            "columnas": len(selected_columns),
            "tiempo": time.time() - started,
            "estructura": representation_label,
            "layout": (
                layout_name.strip()
                or "Mi layout"
            ),
        }

        status.success(
            "Procesamiento completado."
        )

    except Exception as exc:
        st.error(
            "No fue posible completar "
            f"el procesamiento: {exc}"
        )


# ============================================================
# 9. RESULTADO
# ============================================================

if st.session_state.summary:
    result = st.session_state.summary

    st.success(
        "Conversión completada correctamente."
    )

    result_columns = st.columns(5)

    result_columns[0].metric(
        "XML procesados",
        f"{result['procesados']:,}",
    )
    result_columns[1].metric(
        "Correctos",
        f"{result['correctos']:,}",
    )
    result_columns[2].metric(
        "Con error",
        f"{result['errores']:,}",
    )
    result_columns[3].metric(
        "Columnas",
        f"{result['columnas']:,}",
    )
    result_columns[4].metric(
        "Tiempo",
        f"{result['tiempo']:.2f} s",
    )

    st.caption(
        f"Estructura: {result['estructura']} · "
        f"Layout: {result['layout']}"
    )

if st.session_state.download_data:
    st.download_button(
        label=(
            "⬇️ Descargar "
            f"{st.session_state.download_name}"
        ),
        data=st.session_state.download_data,
        file_name=(
            st.session_state.download_name
        ),
        mime=st.session_state.download_mime,
        use_container_width=True,
    )

    st.markdown(
        '<div class="signature">'
        'Aquí debe descargarlo ella 👆<br>'
        'Hecho por tu novio el ingeniero xd 😎'
        '</div>',
        unsafe_allow_html=True,
    )
